from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config.constants import EDITABLE_COLUMNS, MASTER_SHEET_NAME, OUTPUT_DIRECTORY


def update_working_dataframe_from_editor(
    working_master_dataframe: pd.DataFrame,
    edited_section_dataframe: pd.DataFrame,
) -> pd.DataFrame:
    updated_dataframe = working_master_dataframe.copy()

    for row_key, edited_row in edited_section_dataframe.iterrows():
        if str(row_key) == "__total__":
            continue

        row_selector = updated_dataframe["__row_key"] == str(row_key)
        if not row_selector.any():
            continue
        for editable_column in EDITABLE_COLUMNS:
            if editable_column in edited_row.index and editable_column in updated_dataframe.columns:
                updated_dataframe.loc[row_selector, editable_column] = edited_row[editable_column]

    return updated_dataframe.sort_values("__row_order").reset_index(drop=True)


def write_dataframe_back_to_excel(
    updated_master_dataframe: pd.DataFrame,
    workbook_path: str | Path,
    master_sheet_name: str = MASTER_SHEET_NAME,
    create_archive_copy: bool = True,
) -> tuple[Path, Path | None]:
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path)
    worksheet = workbook[master_sheet_name]

    header_lookup = {
        str(worksheet.cell(1, column_index).value).strip(): column_index
        for column_index in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, column_index).value is not None
    }

    for _, row in updated_master_dataframe.iterrows():
        excel_row_number = row.get("__excel_row_number")
        if pd.isna(excel_row_number):
            continue
        excel_row_number = int(excel_row_number)
        for editable_column in EDITABLE_COLUMNS:
            if editable_column not in header_lookup:
                continue
            worksheet.cell(excel_row_number, header_lookup[editable_column]).value = row.get(editable_column)

    workbook.save(workbook_path)

    frozen_copy_path = None
    if create_archive_copy:
        output_directory = workbook_path.parent.parent / OUTPUT_DIRECTORY
        output_directory.mkdir(parents=True, exist_ok=True)
        timestamp_text = datetime.now().strftime("%Y%m%d_%H%M%S")
        frozen_copy_path = output_directory / f"Anjar_manning_frozen_{timestamp_text}.xlsx"
        workbook.save(frozen_copy_path)

    workbook.close()
    return workbook_path, frozen_copy_path


def create_download_workbook_copy(
    updated_master_dataframe: pd.DataFrame,
    source_workbook_path: str | Path,
    master_sheet_name: str = MASTER_SHEET_NAME,
) -> Path:
    source_workbook_path = Path(source_workbook_path)
    output_directory = source_workbook_path.parent.parent / OUTPUT_DIRECTORY
    output_directory.mkdir(parents=True, exist_ok=True)
    download_copy_path = output_directory / "Anjar_manning_download.xlsx"

    workbook = load_workbook(source_workbook_path)
    worksheet = workbook[master_sheet_name]
    header_lookup = {
        str(worksheet.cell(1, column_index).value).strip(): column_index
        for column_index in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, column_index).value is not None
    }

    for _, row in updated_master_dataframe.iterrows():
        excel_row_number = row.get("__excel_row_number")
        if pd.isna(excel_row_number):
            continue
        excel_row_number = int(excel_row_number)
        for editable_column in EDITABLE_COLUMNS:
            if editable_column not in header_lookup:
                continue
            worksheet.cell(excel_row_number, header_lookup[editable_column]).value = row.get(editable_column)

    workbook.save(download_copy_path)
    workbook.close()
    return download_copy_path
