from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config.constants import INPUT_WORKBOOK_PATH, MASTER_SHEET_NAME, WORKING_DIRECTORY, WORKING_WORKBOOK_FILENAME
from config.workbook_config import BUSINESS_VALUE_NORMALIZATION
from core.dataframe_utils import create_row_metadata, normalize_business_values, normalize_column_names


def resolve_input_workbook_path(workbook_path: str | None = None) -> Path:
    selected_path = workbook_path or INPUT_WORKBOOK_PATH
    return Path(selected_path)


def resolve_working_workbook_path(input_workbook_path: str | Path | None = None) -> Path:
    input_path = resolve_input_workbook_path(str(input_workbook_path) if input_workbook_path else None)
    return input_path.parent.parent / WORKING_DIRECTORY / WORKING_WORKBOOK_FILENAME


def workbook_exists(workbook_path: str | Path | None = None) -> bool:
    return resolve_input_workbook_path(str(workbook_path) if workbook_path else None).exists()


def _normalize_business_labels_in_workbook(workbook_path: Path, master_sheet_name: str = MASTER_SHEET_NAME) -> None:
    workbook = load_workbook(workbook_path)
    worksheet = workbook[master_sheet_name]

    header_lookup = {
        str(worksheet.cell(1, column_index).value).strip(): column_index
        for column_index in range(1, worksheet.max_column + 1)
        if worksheet.cell(1, column_index).value is not None
    }
    business_column = header_lookup.get("Business")
    if business_column is None:
        workbook.close()
        return

    for row_index in range(2, worksheet.max_row + 1):
        current_value = worksheet.cell(row_index, business_column).value
        if current_value in BUSINESS_VALUE_NORMALIZATION:
            worksheet.cell(row_index, business_column).value = BUSINESS_VALUE_NORMALIZATION[current_value]

    workbook.save(workbook_path)
    workbook.close()


def _copy_source_to_working(source_path: Path, working_path: Path) -> None:
    working_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, working_path)
    _normalize_business_labels_in_workbook(working_path)


def ensure_working_workbook(input_workbook_path: str | Path | None = None) -> Path:
    source_path = resolve_input_workbook_path(str(input_workbook_path) if input_workbook_path else None)
    working_path = resolve_working_workbook_path(source_path)

    if not working_path.exists():
        _copy_source_to_working(source_path, working_path)
        return working_path

    source_stats = source_path.stat()
    working_stats = working_path.stat()
    if source_stats.st_mtime > working_stats.st_mtime or source_stats.st_size != working_stats.st_size:
        _copy_source_to_working(source_path, working_path)

    return working_path


def reset_working_workbook(input_workbook_path: str | Path | None = None) -> Path:
    source_path = resolve_input_workbook_path(str(input_workbook_path) if input_workbook_path else None)
    working_path = resolve_working_workbook_path(source_path)
    _copy_source_to_working(source_path, working_path)
    return working_path


def load_sheet_dataframe(workbook_path: str | Path, sheet_name: str) -> pd.DataFrame:
    sheet_dataframe = pd.read_excel(workbook_path, sheet_name=sheet_name)
    sheet_dataframe = normalize_column_names(sheet_dataframe)
    sheet_dataframe = normalize_business_values(sheet_dataframe)
    return sheet_dataframe


def load_master_dataframe(workbook_path: str | Path, sheet_name: str) -> pd.DataFrame:
    master_dataframe = load_sheet_dataframe(workbook_path, sheet_name)
    return create_row_metadata(master_dataframe)


def load_workbook_sheet_names(workbook_path: str | Path) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def load_helper_dataframes(workbook_path: str | Path, helper_sheet_names: list[str]) -> dict[str, pd.DataFrame]:
    helper_dataframes: dict[str, pd.DataFrame] = {}
    for helper_sheet_name in helper_sheet_names:
        helper_dataframes[helper_sheet_name] = load_sheet_dataframe(workbook_path, helper_sheet_name)
    return helper_dataframes
